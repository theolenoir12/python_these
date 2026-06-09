# -*- coding: utf-8 -*-
"""
Recherche d'hyperparamètres LSTM pour prédiction de série temporelle.
Sélection du meilleur modèle basée sur le RMSE en prédiction récursive.
Correction du data leakage et de l'alignement des séquences Train/Val.
"""

import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'

import tensorflow as tf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from tensorflow.keras.layers import Input, LSTM, Dense
from tensorflow.keras.callbacks import EarlyStopping
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_absolute_error
import itertools
import csv

# =============================================================================
# CONFIGURATION : grille des hyperparamètres à explorer
# =============================================================================
HYPERPARAMS_GRID = {
    'learning_rate': [1e-3],
    'window_size':   [100],
    'batch_size':    [64],
    'lstm_units':    [100],
    'epochs':        [1000],
}

OUTPUT_DIR = 'resultats_hyperparams'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =============================================================================
# 1. CHARGEMENT, SÉPARATION ET NORMALISATION (SANS FUITE DE DONNÉES)
# =============================================================================
df = pd.read_csv('SoH_ely.csv', sep=';', header=None)
datapro = df.iloc[:, 0].values.reshape(-1, 1)[::24]

n         = len(datapro)
train_end = int(n * 0.70)
val_end   = int(n * 0.85)

# Fit du scaler UNIQUEMENT sur l'entraînement pour éviter la fuite de données futures
scaler = MinMaxScaler(feature_range=(0, 1))
scaler.fit(datapro[:train_end])

# Transformation de tout le dataset
data_normalized = scaler.transform(datapro)

# Récupération des données d'origine (non normalisées) pour l'affichage de la Figure 1
train_data_raw = datapro[:train_end]
val_data_raw   = datapro[train_end:val_end]
test_data_raw  = datapro[val_end:]

print(f"Total: {n} | Train: {len(train_data_raw)} | Val: {len(val_data_raw)} | Test: {len(test_data_raw)}")

# =============================================================================
# 2. UTILITAIRES
# =============================================================================

def create_sequences(data, window_size):
    """Crée les séquences d'entrée X et les cibles y."""
    X, y = [], []
    for i in range(len(data) - window_size):
        X.append(data[i:i + window_size])
        y.append(data[i + window_size])
    if len(X) == 0:
        return np.empty((0, window_size, 1)), np.empty((0,))
    return np.array(X).reshape(-1, window_size, 1), np.array(y)

def predict_recursive(model, seed_window, n_steps):
    """
    Prédiction récursive : réutilise chaque prédiction comme entrée suivante.
    seed_window : (window_size, 1) — les ws derniers points connus.
    """
    window = seed_window.copy().flatten().tolist()
    predictions = []
    ws = len(window)
    for _ in range(n_steps):
        x_in = np.array(window[-ws:]).reshape(1, ws, 1)
        pred = model.predict(x_in, verbose=0)[0, 0]
        predictions.append(pred)
        window.append(pred)
    return np.array(predictions).reshape(-1, 1)

def compute_metrics_from_arrays(y_true_real, y_pred_real):
    mse  = float(np.mean((y_pred_real - y_true_real)**2))
    rmse = float(np.sqrt(mse))
    mae  = float(mean_absolute_error(y_true_real, y_pred_real))
    return {'mse': mse, 'rmse': rmse, 'mae': mae}

def compute_metrics(model, X, y_true_norm, scaler):
    y_pred_norm = model.predict(X, verbose=0)
    mse_norm    = float(np.mean((y_pred_norm - y_true_norm.reshape(-1, 1))**2))
    y_pred_real = scaler.inverse_transform(y_pred_norm)
    y_true_real = scaler.inverse_transform(y_true_norm.reshape(-1, 1))
    rmse = float(np.sqrt(np.mean((y_pred_real - y_true_real)**2)))
    mae  = float(mean_absolute_error(y_true_real, y_pred_real))
    return {'mse_norm': mse_norm, 'rmse': rmse, 'mae': mae}, y_pred_real, y_true_real

def build_model(window_size, lstm_units, learning_rate):
    model = tf.keras.Sequential([
        Input(shape=(window_size, 1)),
        LSTM(lstm_units),
        Dense(1)
    ])
    model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=learning_rate), loss='mse')
    return model

# =============================================================================
# 3. FIGURE RÉPARTITION DES DONNÉES
# =============================================================================
fig_split, ax = plt.subplots(figsize=(12, 4))
ax.plot(range(train_end),          train_data_raw, label='Entraînement (70%)', color='steelblue')
ax.plot(range(train_end, val_end), val_data_raw,   label='Validation (15%)',   color='darkorange')
ax.plot(range(val_end, n),         test_data_raw,  label='Test (15%)',         color='seagreen')
ax.set_title("Répartition des données (SoH réel)")
ax.set_xlabel("Index temporel (jours)")
ax.set_ylabel("SoH")
ax.legend(); ax.grid(True)
fig_split.tight_layout()
fig_split.savefig(os.path.join(OUTPUT_DIR, 'repartition_donnees.png'), dpi=150)
plt.close(fig_split)
print("Figure répartition enregistrée.")

# =============================================================================
# 4. BOUCLE DE RECHERCHE D'HYPERPARAMÈTRES
# =============================================================================
keys   = list(HYPERPARAMS_GRID.keys())
combos = list(itertools.product(*[HYPERPARAMS_GRID[k] for k in keys]))
total  = len(combos)

results_log       = []
skipped_log       = []  
best_val_rmse_rec = np.inf  # On cherche désormais à minimiser le RMSE récursif
best_combo        = None

print(f"\n=== Début de la recherche : {total} combinaisons ===\n")

for idx, values in enumerate(combos, start=1):
    tf.keras.backend.clear_session() # Nettoyage mémoire RAM en début de boucle
    
    hp = dict(zip(keys, values))
    lr, ws, bs, units, max_epochs = (hp['learning_rate'], hp['window_size'],
                                     hp['batch_size'],    hp['lstm_units'],
                                     hp['epochs'])
    run_id = f"lr{lr}_ws{ws}_bs{bs}_u{units}_ep{max_epochs}"
    print(f"[{idx}/{total}] {run_id}")

    # -- Préparation des séquences avec Overlap --
    # Train normal
    train_slice = data_normalized[:train_end]
    
    # Pour la validation, on inclut les 'ws' derniers points de l'entraînement
    # Ainsi, le 1er y_val généré correspondra exactement à data_normalized[train_end]
    val_slice = data_normalized[train_end - ws : val_end]

    X_train, y_train = create_sequences(train_slice, ws)
    X_val,   y_val   = create_sequences(val_slice,   ws)

    if len(X_train) == 0 or len(X_val) == 0:
        msg = f"SKIP — séquences vides ou window_size trop grande"
        print(f"   {msg}"); skipped_log.append({'run_id': run_id, 'raison': msg}); continue

    # -- Entraînement --
    model = build_model(ws, units, lr)
    early_stop = EarlyStopping(monitor='val_loss', patience=50, restore_best_weights=True, verbose=0)
    history = model.fit(X_train, y_train, epochs=max_epochs, batch_size=bs,
                        validation_data=(X_val, y_val), callbacks=[early_stop], verbose=0)

    if 'val_loss' not in history.history:
        msg = "SKIP — val_loss absent de l'historique"
        print(f"   {msg}"); skipped_log.append({'run_id': run_id, 'raison': msg}); continue

    epochs_run       = len(history.history['loss'])
    best_val_mse_run = min(history.history['val_loss'])

    # -- Métriques teacher-forcing (one-step-ahead) --
    train_metrics, y_train_pred, y_train_true = compute_metrics(model, X_train, y_train, scaler)

    # -- Prédiction récursive sur la validation --
    # La graine est exactement les ws derniers points de l'entraînement (soit X_val[0])
    seed = train_slice[-ws:] 
    n_val_steps = len(y_val) # Correspond exactement à val_end - train_end
    
    y_val_pred_rec_norm = predict_recursive(model, seed, n_val_steps)
    y_val_pred_rec      = scaler.inverse_transform(y_val_pred_rec_norm)
    y_val_true_rec      = scaler.inverse_transform(y_val.reshape(-1, 1))
    
    rec_metrics = compute_metrics_from_arrays(y_val_true_rec, y_val_pred_rec)

    print(
        f"   Epochs: {epochs_run}/{max_epochs} | "
        f"Train MSE={train_metrics['mse_norm']:.5f} | "
        f"Val RMSE(rec)={rec_metrics['rmse']:.4f} | "
        f"Val MAE(rec)={rec_metrics['mae']:.4f}"
    )

    # -- Log --
    row = {
        'run_id': run_id, 'learning_rate': lr, 'window_size': ws,
        'batch_size': bs, 'lstm_units': units, 'max_epochs': max_epochs,
        'epochs_run': epochs_run,
        'train_mse_norm': round(train_metrics['mse_norm'], 6),
        'train_rmse':     round(train_metrics['rmse'],     6),
        'train_mae':      round(train_metrics['mae'],      6),
        'val_rmse_rec':   round(rec_metrics['rmse'],       6),
        'val_mae_rec':    round(rec_metrics['mae'],        6),
    }
    results_log.append(row)

    # ===================================================================
    # FIGURE 1 : Courbes de loss
    # ===================================================================
    fig = plt.figure(figsize=(14, 6))
    gs  = gridspec.GridSpec(1, 2, width_ratios=[2, 1])

    ax_loss = fig.add_subplot(gs[0])
    ax_loss.semilogy(history.history['loss'],     label='Loss entraînement (MSE)', color='steelblue')
    ax_loss.semilogy(history.history['val_loss'], label='Loss validation (MSE)',   color='darkorange')
    ax_loss.set_title(f"Courbes de Loss (log Y) — {run_id}", fontsize=10)
    ax_loss.set_xlabel("Epochs")
    ax_loss.set_ylabel("MSE (normalisé, échelle log)")
    ax_loss.legend(); ax_loss.grid(True, which='both', linestyle='--', alpha=0.6)

    best_ep  = int(np.argmin(history.history['val_loss']))
    best_val = history.history['val_loss'][best_ep]
    ax_loss.annotate(
        f"min={best_val:.5f}\n(ep {best_ep+1})",
        xy=(best_ep, best_val),
        xytext=(best_ep + max(1, epochs_run * 0.08), best_val * 3),
        arrowprops=dict(arrowstyle='->', color='red'), fontsize=8, color='red'
    )

    ax_tab = fig.add_subplot(gs[1])
    ax_tab.axis('off')
    table_data = [
        ['Hyperparamètre',   'Valeur'],
        ['learning_rate',    str(lr)],
        ['window_size',      str(ws)],
        ['batch_size',       str(bs)],
        ['lstm_units',       str(units)],
        ['epochs réalisés',  f"{epochs_run}/{max_epochs}"],
        ['', ''],
        ['Métrique',         'Valeur'],
        ['Train RMSE',       f"{train_metrics['rmse']:.4f}"],
        ['Val RMSE (rec)',   f"{rec_metrics['rmse']:.4f}"],
        ['Val MAE  (rec)',   f"{rec_metrics['mae']:.4f}"],
    ]
    tbl = ax_tab.table(cellText=table_data, loc='center', cellLoc='left')
    tbl.auto_set_font_size(False); tbl.set_fontsize(8); tbl.scale(1.15, 1.35)
    for col in range(2):
        for hr in [0, 7]:
            tbl[(hr, col)].set_facecolor('#2c3e50')
            tbl[(hr, col)].set_text_props(color='white', fontweight='bold')

    fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, f'loss_{run_id}.png'), dpi=150)
    # plt.close(fig)

    # ===================================================================
    # FIGURE 2 : Comparaison Réel vs Prédit (Alignement Corrigé)
    # ===================================================================
    fig2, axes2 = plt.subplots(2, 1, figsize=(14, 13), sharex=False)

    # Les index sont maintenant parfaitement alignés grâce à la correction
    x_train = np.arange(ws, train_end)
    x_val   = np.arange(train_end, val_end)

    # --- Entraînement ---
    axes2[0].plot(x_train, y_train_true,  label='Réel',        color='steelblue',  linewidth=0.9)
    axes2[0].plot(x_train, y_train_pred, label='Prédit', color='tomato', linestyle='--', linewidth=0.9)
    axes2[0].set_title(f"Entraînement  — RMSE={train_metrics['rmse']:.4f}", fontsize=10)
    axes2[0].set_ylabel("SoH"); axes2[0].legend(fontsize=8); axes2[0].grid(True)
    axes2[0].axvline(train_end, color='gray', linestyle=':', linewidth=0.8)

    # --- Validation récursive ---
    axes2[1].plot(x_val, y_val_true_rec,   label='Réel',              color='darkorange', linewidth=0.9)
    axes2[1].plot(x_val, y_val_pred_rec,   label='Prédit (récursif)', color='purple',     linestyle='--', linewidth=0.9)
    axes2[1].set_title(f"Validation récursive (autorégressif) — RMSE={rec_metrics['rmse']:.4f}", fontsize=10)
    axes2[1].set_xlabel("Temps (jours depuis début du dataset)")
    axes2[1].set_ylabel("SoH"); axes2[1].legend(fontsize=8); axes2[1    ].grid(True)
    
    # On aligne l'axe X des graphiques de validation pour comparer facilement
    axes2[1].set_xlim(axes2[1].get_xlim())

    fig2.suptitle(f"Comparaison Réel vs Prédit — {run_id}", fontsize=11, fontweight='bold')
    fig2.tight_layout()
    fig2.savefig(os.path.join(OUTPUT_DIR, f'comparaison_{run_id}.png'), dpi=150)
    # plt.close(fig2)
    plt.show()

    # -- SAUVEGARDE DU MEILLEUR MODÈLE (Basé sur le RMSE Récursif) --
    if rec_metrics['rmse'] < best_val_rmse_rec:
        best_val_rmse_rec = rec_metrics['rmse']
        best_combo        = hp.copy()
        model.save(os.path.join(OUTPUT_DIR, 'meilleur_modele.keras'))
        print(f"   *** Nouveau meilleur modèle (Val RMSE rec={best_val_rmse_rec:.5f}) ***")

# =============================================================================
# 5. EXPORT CSV
# =============================================================================
if results_log:
    csv_path = os.path.join(OUTPUT_DIR, 'resultats_hyperparams2.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=results_log[0].keys())
        writer.writeheader(); writer.writerows(results_log)
    print(f"\nCSV enregistré : {csv_path}")

if skipped_log:
    skip_path = os.path.join(OUTPUT_DIR, 'combinaisons_ignorees.csv')
    with open(skip_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['run_id', 'raison'])
        writer.writeheader(); writer.writerows(skipped_log)

# =============================================================================
# 6. EXPORT EXCEL RÉCAPITULATIF
# =============================================================================
if results_log:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule

        wb = openpyxl.Workbook()
        ws_sheet = wb.active
        ws_sheet.title = "Résultats"

        # --- Styles ---
        DARK_HEADER = "2C3E50"
        BLUE = "1A5276"
        PURPLE = "6E2F77"
        BEST_ROW = "FDEBD0"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws_sheet.merge_cells('A1:O1')
        c = ws_sheet['A1']
        c.value = "Recherche d'hyperparamètres LSTM — Résultats"
        c.font = Font(bold=True, size=13, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK_HEADER)
        c.alignment = center
        ws_sheet.row_dimensions[1].height = 28

        groups = [
            ('A', 'G', "Hyperparamètres", DARK_HEADER),
            ('H', 'J', "Entraînement", BLUE),
            ('K', 'L', "Validation récursive", PURPLE),
        ]

        for sc, ec, label, color in groups:
            ws_sheet.merge_cells(f"{sc}2:{ec}2")
            c = ws_sheet[f"{sc}2"]
            c.value = label
            c.font = header_font
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = center

        columns = [
            ('run_id', "Run ID"), ('learning_rate', "Learning Rate"),
            ('window_size', "Window Size"), ('batch_size', "Batch Size"),
            ('lstm_units', "LSTM Units"), ('max_epochs', "Epochs (max)"),
            ('epochs_run', "Epochs réalisés"),
            ('train_mse_norm', "Train MSE"), ('train_rmse', "Train RMSE"), ('train_mae', "Train MAE"),
            ('val_rmse_rec', "Val RMSE (rec)"), ('val_mae_rec', "Val MAE (rec)"),
        ]

        header_colors = [DARK_HEADER]*7 + [BLUE]*3 + [PURPLE]*3 + [BLUE]*2

        for col_i, ((key, label), color) in enumerate(zip(columns, header_colors), start=1):
            c = ws_sheet.cell(row=3, column=col_i, value=label)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = center
            c.border = thin_border

        ws_sheet.row_dimensions[3].height = 28

        best_run_id = None
        if best_combo:
            best_run_id = (f"lr{best_combo['learning_rate']}_ws{best_combo['window_size']}_"
                           f"bs{best_combo['batch_size']}_u{best_combo['lstm_units']}_ep{best_combo['epochs']}")

        for row_i, row in enumerate(results_log, start=4):
            is_best = (row['run_id'] == best_run_id)
            for col_i, (key, _) in enumerate(columns, start=1):
                val = row.get(key, None)
                c = ws_sheet.cell(row=row_i, column=col_i, value=val)
                c.alignment = center
                c.border = thin_border
                if is_best: c.fill = PatternFill("solid", fgColor=BEST_ROW)
                if col_i >= 11: c.font = Font(bold=True)

        # Règle de couleur basée sur la validation récursive (Colonne N)
        val_rmse_rec_col = get_column_letter(14)
        ws_sheet.conditional_formatting.add(
            f"{val_rmse_rec_col}4:{val_rmse_rec_col}{3+len(results_log)}",
            ColorScaleRule(start_type='min', start_color='63BE7B', end_type='max', end_color='F8696B')
        )

        widths = [28,14,13,11,11,12,15,14,10,10,14,10,10,12,12]
        for i, w in enumerate(widths, start=1):
            ws_sheet.column_dimensions[get_column_letter(i)].width = w

        ws_sheet.freeze_panes = "A4"
        xl_path = os.path.join(OUTPUT_DIR, "resultats_hyperparams2.xlsx")
        wb.save(xl_path)
        print(f"Excel enregistré : {xl_path}")

    except ImportError:
        print("openpyxl non installé → pip install openpyxl")

# =============================================================================
# 7. FIGURE RÉCAPITULATIVE (Mise à jour pour inclure la perf récursive)
# =============================================================================
if results_log:
    run_ids      = [r['run_id']       for r in results_log]
    val_rmses_rc = [r['val_rmse_rec'] for r in results_log]
    val_maes_rc  = [r['val_mae_rec']  for r in results_log]

    x = np.arange(len(run_ids))
    fig_recap, axes = plt.subplots(3, 1, figsize=(max(12, len(run_ids) * 1.5), 10))

    for ax_r, vals, ylabel, title, color in zip(
        axes,
        [val_rmses_rc, val_maes_rc],
        ['RMSE', 'MAE'],
        ['Validation RMSE (Récursif - Critère)', 'Validation MAE (Récursif)'],
        ['purple', 'seagreen']
    ):
        ax_r.bar(x, vals, color=color)
        ax_r.set_title(title)
        ax_r.set_ylabel(ylabel)
        ax_r.set_xticks(x)
        ax_r.set_xticklabels(run_ids, rotation=45, ha='right', fontsize=7)
        ax_r.grid(axis='y')

    # Mise en évidence du meilleur basé sur le RMSE récursif
    best_idx = int(np.argmin(val_rmses_rc))
    for ax_r, vals in zip(axes, [val_rmses_rc, val_maes_rc]):
        ax_r.patches[best_idx].set_facecolor('crimson')
        ax_r.annotate('Meilleur', xy=(best_idx, vals[best_idx]),
                      xytext=(best_idx, vals[best_idx] * 1.05),
                      ha='center', fontsize=8, color='crimson', fontweight='bold')

    fig_recap.suptitle("Comparaison de toutes les combinaisons d'hyperparamètres", fontweight='bold')
    fig_recap.tight_layout()
    fig_recap.savefig(os.path.join(OUTPUT_DIR, 'recap_toutes_combinaisons.png'), dpi=150)
    plt.close(fig_recap)

# =============================================================================
# 8. RÉSUMÉ FINAL
# =============================================================================
print("\n" + "="*60)
if best_combo:
    print("MEILLEURE COMBINAISON D'HYPERPARAMÈTRES (Basée sur le RMSE Récursif)")
    print("="*60)
    for k, v in best_combo.items():
        print(f"  {k:20s} : {v}")
    print(f"  {'val_rmse_rec':20s} : {best_val_rmse_rec:.6f}")
    print(f"\nModèle sauvegardé : {os.path.join(OUTPUT_DIR, 'meilleur_modele.keras')}")
else:
    print("Aucune combinaison valide — aucun modèle sauvegardé.")
print(f"Figures et CSV/Excel : {OUTPUT_DIR}/")