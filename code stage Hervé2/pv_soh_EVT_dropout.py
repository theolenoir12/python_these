# -*- coding: utf-8 -*-
"""
Recherche d'hyperparamètres LSTM avec balayage sur le Dropout.
Sélection du meilleur modèle basée sur le RMSE en prédiction récursive.
"""

import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'

import tensorflow as tf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from tensorflow.keras.layers import Input, LSTM, Dense, Dropout
from tensorflow.keras.callbacks import EarlyStopping
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_absolute_error
import itertools
import csv

# =============================================================================
# CONFIGURATION : grille des hyperparamètres (8 simulations au total)
# =============================================================================
HYPERPARAMS_GRID = {
    'learning_rate': [1e-3],
    'window_size':   [100],
    'batch_size':    [64],
    'lstm_units':    [50, 100],  # 2 valeurs
    'dropout':       [0.0, 0.1, 0.2, 0.3], # 4 valeurs -> 2 * 4 = 8 sims
    'epochs':        [1000],
}

OUTPUT_DIR = 'resultats_hyperparams'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =============================================================================
# 1. CHARGEMENT ET PRÉPARATION
# =============================================================================
df = pd.read_csv('SoH_ely.csv', sep=';', header=None)
datapro = df.iloc[:, 0].values.reshape(-1, 1)[::24]

n         = len(datapro)
train_end = int(n * 0.70)
val_end   = int(n * 0.85)

scaler = MinMaxScaler(feature_range=(0, 1))
scaler.fit(datapro[:train_end])
data_normalized = scaler.transform(datapro)

train_data_raw = datapro[:train_end]
val_data_raw   = datapro[train_end:val_end]
test_data_raw  = datapro[val_end:]

# =============================================================================
# 2. UTILITAIRES
# =============================================================================

def create_sequences(data, window_size):
    X, y = [], []
    for i in range(len(data) - window_size):
        X.append(data[i:i + window_size])
        y.append(data[i + window_size])
    if len(X) == 0: return np.empty((0, window_size, 1)), np.empty((0,))
    return np.array(X).reshape(-1, window_size, 1), np.array(y)

def predict_recursive(model, seed_window, n_steps):
    window = seed_window.copy().flatten().tolist()
    predictions = []
    ws = len(window)
    for _ in range(n_steps):
        x_in = np.array(window[-ws:]).reshape(1, ws, 1)
        pred = model.predict(x_in, verbose=0)[0, 0]
        predictions.append(pred)
        window.append(pred)
    return np.array(predictions).reshape(-1, 1)

def compute_metrics_from_arrays(y_true_real, y_pred_real):
    mse  = float(np.mean((y_pred_real - y_true_real)**2))
    rmse = float(np.sqrt(mse))
    mae  = float(mean_absolute_error(y_true_real, y_pred_real))
    return {'mse': mse, 'rmse': rmse, 'mae': mae}

def compute_metrics(model, X, y_true_norm, scaler):
    y_pred_norm = model.predict(X, verbose=0)
    mse_norm    = float(np.mean((y_pred_norm - y_true_norm.reshape(-1, 1))**2))
    y_pred_real = scaler.inverse_transform(y_pred_norm)
    y_true_real = scaler.inverse_transform(y_true_norm.reshape(-1, 1))
    rmse = float(np.sqrt(np.mean((y_pred_real - y_true_real)**2)))
    mae  = float(mean_absolute_error(y_true_real, y_pred_real))
    return {'mse_norm': mse_norm, 'rmse': rmse, 'mae': mae}, y_pred_real, y_true_real

def build_model(window_size, lstm_units, learning_rate, dropout_rate):
    model = tf.keras.Sequential([
        Input(shape=(window_size, 1)),
        LSTM(lstm_units, dropout=dropout_rate), # Ajout du dropout ici
        Dense(1)
    ])
    model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=learning_rate), loss='mse')
    return model

# =============================================================================
# 4. BOUCLE DE RECHERCHE D'HYPERPARAMÈTRES
# =============================================================================
keys   = list(HYPERPARAMS_GRID.keys())
combos = list(itertools.product(*[HYPERPARAMS_GRID[k] for k in keys]))
total  = len(combos)

results_log       = []
skipped_log       = []  
best_val_rmse_rec = np.inf
best_combo        = None

print(f"\n=== Début de la recherche : {total} combinaisons (Dropout sweep) ===\n")

for idx, values in enumerate(combos, start=1):
    tf.keras.backend.clear_session()
    
    hp = dict(zip(keys, values))
    lr, ws, bs, units, dr, max_epochs = (hp['learning_rate'], hp['window_size'],
                                         hp['batch_size'],    hp['lstm_units'],
                                         hp['dropout'],       hp['epochs'])
    
    run_id = f"lr{lr}_ws{ws}_bs{bs}_u{units}_do{dr}_ep{max_epochs}"
    print(f"[{idx}/{total}] {run_id}")

    train_slice = data_normalized[:train_end]
    val_slice = data_normalized[train_end - ws : val_end]

    X_train, y_train = create_sequences(train_slice, ws)
    X_val,   y_val   = create_sequences(val_slice,   ws)

    if len(X_train) == 0 or len(X_val) == 0:
        continue

    # Construction avec Dropout
    model = build_model(ws, units, lr, dr)
    early_stop = EarlyStopping(monitor='val_loss', patience=200, restore_best_weights=True, verbose=0)
    history = model.fit(X_train, y_train, epochs=max_epochs, batch_size=bs,
                        validation_data=(X_val, y_val), callbacks=[early_stop], verbose=0)

    epochs_run       = len(history.history['loss'])
    train_metrics, y_train_pred, y_train_true = compute_metrics(model, X_train, y_train, scaler)

    # Prédiction récursive
    seed = train_slice[-ws:] 
    n_val_steps = len(y_val)
    y_val_pred_rec_norm = predict_recursive(model, seed, n_val_steps)
    y_val_pred_rec      = scaler.inverse_transform(y_val_pred_rec_norm)
    y_val_true_rec      = scaler.inverse_transform(y_val.reshape(-1, 1))
    rec_metrics = compute_metrics_from_arrays(y_val_true_rec, y_val_pred_rec)

    print(f"    Val RMSE(rec)={rec_metrics['rmse']:.4f} | Epochs: {epochs_run}")

    row = {
        'run_id': run_id, 'learning_rate': lr, 'window_size': ws,
        'batch_size': bs, 'lstm_units': units, 'dropout': dr, 'max_epochs': max_epochs,
        'epochs_run': epochs_run,
        'train_mse_norm': round(train_metrics['mse_norm'], 6),
        'train_rmse':     round(train_metrics['rmse'],     6),
        'val_rmse_rec':   round(rec_metrics['rmse'],       6),
        'val_mae_rec':    round(rec_metrics['mae'],        6),
    }
    results_log.append(row)

    # Sauvegarde du meilleur
    if rec_metrics['rmse'] < best_val_rmse_rec:
        best_val_rmse_rec = rec_metrics['rmse']
        best_combo        = hp.copy()
        model.save(os.path.join(OUTPUT_DIR, 'meilleur_modele_dropout.keras'))

# =============================================================================
# 5. EXPORTS (Fichiers renommés avec suffixe dropout)
# =============================================================================
csv_path = os.path.join(OUTPUT_DIR, 'resultats_hyperparams_dropout.csv')
xl_path  = os.path.join(OUTPUT_DIR, 'resultats_hyperparams_dropout.xlsx')

if results_log:
    # Export CSV
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=results_log[0].keys())
        writer.writeheader(); writer.writerows(results_log)

    # Export Excel Stylisé
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws_sheet = wb.active
        ws_sheet.title = "Résultats Dropout"

        # Entêtes
        columns = list(results_log[0].keys())
        for col_i, label in enumerate(columns, start=1):
            c = ws_sheet.cell(row=1, column=col_i, value=label)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2C3E50")
            c.alignment = Alignment(horizontal='center')

        # Données
        for row_i, data in enumerate(results_log, start=2):
            for col_i, key in enumerate(columns, start=1):
                c = ws_sheet.cell(row=row_i, column=col_i, value=data[key])
                c.alignment = Alignment(horizontal='center')
        
        wb.save(xl_path)
        print(f"\nFichiers sauvegardés :\n - {csv_path}\n - {xl_path}")
    except:
        print("\nCSV sauvegardé. (Openpyxl manquant pour l'Excel stylisé)")

print("\nTerminé.")